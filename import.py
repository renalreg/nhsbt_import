"""
A script for parsing the NHSBT file and importing it into the database. The script
will also create an audit file which will contain all the new and updated patients.
An error log will also be created to aid in debugging.

Typical usage example:
    Both examples will create an error and audit file in the directory specified by the -d flag
    This will run the script without committing the changes to the database
    poetry run import.py -d /path/to/the/directory

    This will run the script and commit the changes to the database
    poetry run import.py -d /path/to/the/directory -c

Args:
    -d (--directory): The directory containing the NHSBT file
    -c (--commit): Commit the changes to the database

Raises:
    ValueError: Number of columns in the NHSBT file isn't as expected

Returns:
    None

"""

import os
import warnings
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy.orm import Session
from ukrr_models.nhsbt_models import UKTPatient, UKTTransplant  # type: ignore [import]
from ukrr_models.rr_models import UKRR_Deleted_Patient  # type: ignore [import]

from nhsbt_import import utils
from nhsbt_import.df_columns import df_columns

warnings.simplefilter(action="ignore", category=FutureWarning)
args = utils.args_parse()
log = utils.create_logs(args.directory)


def import_patient(
    index: int,
    row: pd.Series,
    output_dfs: dict[str, pd.DataFrame],
    session: Session,
) -> Optional[str]:
    """
    Take a patient row and checks to see if the uktssa is present in the database. If not
    match type is set to new, a row is created in the output file and a new patient is
    committed to the database. If the patient does exist a comparison is done to see if
    an update is required. If it is match type is set to update, an entry is added to
    the corresponding output sheet and an update is committed to the database. If
    no update is require match type is set to existing but that is all. If more than
    one entry is found for a patient an error is logged and match type is left as none
    which will skip any attempt to upload transplant data.

    Args:
        index (int): The row number from the NHSBT file. Used for messaging
        row (pd.Series): A row relating to a patient from the NHSBT file
        output_dfs (dict[str, pd.DataFrame]): A dict of dataframes for outputs
        session (Session): An sqlalch session

    Returns:
        Optional[str]: A match type (New, Update, existing)
    """
    match_type = None

    incoming_patient = utils.create_incoming_patient(index, row)
    # If len == 1 patient exists, check if update is required
    if (
        len(
            results := (
                session.query(UKTPatient)
                .filter_by(uktssa_no=incoming_patient.uktssa_no)
                .all()
            )
        )
        == 1
    ):
        log.info("UKT Patient %s found in database", incoming_patient.uktssa_no)
        existing_patient = results[0]

        if utils.compare_patients(incoming_patient, existing_patient):
            match_type = "Existing"
            log.info("No Update required")
        else:
            log.info("Updating patient")

            match_type = "Update"

            match_row = utils.make_patient_match_row(
                match_type, incoming_patient, existing_patient
            )

            output_dfs["updated_patients"] = utils.add_df_row(
                output_dfs["updated_patients"], match_row
            )

            utils.update_nhsbt_patient(incoming_patient, existing_patient)

    # If len == 0 add patient to DB
    elif len(results) == 0:
        log.info("Adding patient %s", incoming_patient.uktssa_no)
        match_type = "New"

        match_row = utils.make_patient_match_row(
            match_type, incoming_patient, existing_patient=None
        )

        output_dfs["new_patients"] = utils.add_df_row(
            output_dfs["new_patients"], match_row
        )

        session.add(incoming_patient)

    # If len > 1 something is wrong, raise
    else:
        log.error("%s in the database multiple times", incoming_patient.uktssa_no)

    return match_type


def import_transplants(
    index: int, row: pd.Series, output_dfs: dict[str, pd.DataFrame], session: Session
) -> list[str]:
    """
    Loops over a row looking for dates of registration on the transplant list which is the
    minimum requirement for a transplant entry. Whenever one is found a transplant object
    is created and as part of that process a registration id is created. This registration id
    is added to the list which is returned as a product of this function. The database is
    consulted to find existing transplants which a matching id. If none is found an
    entry is added with a match type of new to the output and committed to the database.
    If one is found the two are compared to see if an update is required. If it is, an entry
    is added to the output and a commit is made to the database. If more than one entry
    is found in the database an error is logged.

    max_transplants is determined by what is sent in the file and will need to be adjusted
    if more columns of transplants are sent. Currently the max is 6

    Args:
        row (pd.Series): A row relating to a patient from the NHSBT file
        output_dfs (dict[str, pd.DataFrame]): A dict of dataframes for outputs
        session (Session): An sqlalch session

    Returns:
        list[int]: A list of all the registration IDs for transplants
    """

    ##################
    max_transplants = 6
    ##################

    transplant_counter = 1
    registration_ids = []

    while (
        transplant_counter <= max_transplants
        and row[f"uktr_date_on{transplant_counter}"] != ""
    ):
        incoming_transplant = utils.create_incoming_transplant(
            index, row, transplant_counter
        )
        registration_ids.append(incoming_transplant.registration_id)
        # If len == 1 transplant exists, check if update is required
        results = (
            session.query(UKTTransplant)
            .filter_by(registration_id=incoming_transplant.registration_id)
            .all()
        )
        if len(results) == 1:
            log.info(
                "Registration ID %s found in database",
                incoming_transplant.registration_id,
            )

            existing_transplant = results[0]

            if utils.compare_transplants(incoming_transplant, existing_transplant):
                log.info("No Update required")
            else:
                log.info("Updating transplant")

                match_row = utils.make_transplant_match_row(
                    "Update", incoming_transplant, existing_transplant
                )

                output_dfs["updated_transplants"] = utils.add_df_row(
                    output_dfs["updated_transplants"], match_row
                )

                utils.update_nhsbt_transplant(incoming_transplant, existing_transplant)

        # If len == 0 add transplant to DB
        elif len(results) == 0:
            log.info("Adding transplant %s", incoming_transplant.registration_id)

            match_row = utils.make_transplant_match_row(
                "New", incoming_transplant, existing_transplant=None
            )

            output_dfs["new_transplants"] = utils.add_df_row(
                output_dfs["new_transplants"], match_row
            )

            session.add(incoming_transplant)

        # If len > 1 something is wrong, raise
        else:
            log.error(
                "%s in the database multiple times", incoming_transplant.registration_id
            )
        transplant_counter += 1

    return registration_ids


def nhsbt_import(input_file_path: str, audit_file_path: str, session: Session):
    # THIS IS NOW BREAKING PYLINT BECAUSE IT'S TOO LONG
    """
    Reads in the NHSBT file and builds all the output dataframes. Uses import_patients()
    and import_transplants to import the data to the database and build the out puts. Runs
    check on all patients and transplants to make sure nothing is missing from the file that
    was previously included and checks against the deleted patients table to make sure no
    patients have been deleted in error.

    Expected number of columns will need to be adjusted if NHSBT change the shape of their.
    If we get more or less an error is raised.

    Args:
        input_file_path (str): NHSBT file path
        audit_file_path (str): Output file path
        session (Session): An sqlalch session

    Raises:
        ValueError: Number of columns in the NHSBT file isn't as expected
    """

    ###################################
    expected_number_of_columns = 125
    ###################################

    nhsbt_df = pd.read_csv(
        input_file_path,
        na_filter=False,
        skip_blank_lines=True,
    )
    utils.column_is_int(nhsbt_df, "UKTR_ID")

    nhsbt_number_of_columns = nhsbt_df.shape[1]

    if expected_number_of_columns != nhsbt_number_of_columns:
        raise ValueError(
            f"""
            Expected {expected_number_of_columns} columns in the NHSBT file
            There are {nhsbt_number_of_columns}
            """
        )

    output_dfs = utils.create_output_dfs(df_columns)
    registration_ids = []

    for index, row in nhsbt_df.iterrows():
        index += 1  # type: ignore [operator]
        log.info("on line %s", index + 1)
        if import_patient(index, row, output_dfs, session):
            registration_ids.extend(import_transplants(index, row, output_dfs, session))

    file_uktssas = nhsbt_df["UKTR_ID"].tolist()

    if missing_uktssa := utils.check_missing_patients(session, file_uktssas):
        missing_patients = batch_query(
            missing_uktssa, session, UKTPatient, UKTPatient.uktssa_no
        )

        patient_data = pd.DataFrame(
            [
                utils.make_missing_patient_row("Missing", missing_patient)
                for missing_patient in missing_patients
            ]
        )

        output_dfs["missing_patients"] = pd.concat(
            [output_dfs["missing_patients"], patient_data], axis=0, ignore_index=True
        )  # type: ignore [operator]

    if missing_transplants_ids := utils.check_missing_transplants(
        session, registration_ids
    ):
        missing_transplants = batch_query(
            missing_transplants_ids,
            session,
            UKTTransplant,
            UKTTransplant.registration_id,
        )

        transplant_dataframe = pd.DataFrame(
            [
                utils.make_missing_transplant_match_row(missing_transplant)
                for missing_transplant in missing_transplants
            ]
        )

        output_dfs["missing_transplants"] = pd.concat(
            [output_dfs["missing_transplants"], transplant_dataframe],
            axis=0,
            ignore_index=True,
        )

    if deleted_uktssa := utils.deleted_patient_check(session, file_uktssas):
        deleted_patients = batch_query(
            deleted_uktssa,
            session,
            UKRR_Deleted_Patient,
            UKRR_Deleted_Patient.uktssa_no,
        )

        deleted_data = pd.DataFrame(
            utils.make_deleted_patient_row("Deleted", deleted_patient)
            for deleted_patient in deleted_patients
        )

        output_dfs["deleted_patients"] = pd.concat(
            [output_dfs["deleted_patients"], deleted_data], axis=0, ignore_index=True
        )

    wb = Workbook()
    wb.remove(wb["Sheet"])

    for sheet_name, df in output_dfs.items():
        if df.empty:
            continue
        ws = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for i, col in enumerate(df.columns):
            column = get_column_letter(i + 1)
            header_length = len(col)
            cell_length = max(
                df[col].astype(str).map(len).astype(int).max(), header_length
            )
            adjusted_width = int(cell_length) + 2
            ws.column_dimensions[column].width = adjusted_width

        for row in ws:
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

    if len(wb.sheetnames) == 0:
        log.info("Nothing to write to audit file")
    else:
        if "updated_patients" in wb.sheetnames:
            utils.colour_differences(wb, "updated_patients")
        if "updated_transplants" in wb.sheetnames:
            utils.colour_differences(wb, "updated_transplants")
        wb.save(audit_file_path)


def batch_query(keys, session, query, key_filter):
    """
    Batch query function
    Args:
        keys: list of values to filter for
        session: sqlachemy session
        query: sqlachemy orm to query
        key_filter: sqlachemy orm column to filter on

    Returns:
        results: list of query results
    """
    results = []
    batch_size = 1000
    for i in range(0, len(keys), batch_size):
        batch = keys[i : i + batch_size]
        batch_results = session.query(query).filter(key_filter.in_(batch)).all()
        results.extend(batch_results)
    return results


def main():
    """
    Main function for the script. Creates a session, gets the input file path, creates
    the audit file path and runs the import function. If the commit
    flag is set the session is committed and closed. If not the session is closed without
    committing.
    """
    input_file_path = utils.get_input_file_path(args.directory)
    audit_file_path = os.path.join(args.directory, "audit.xlsx")
    utils.clean_csv(input_file_path)
    session = utils.create_session()
    nhsbt_import(input_file_path, audit_file_path, session)
    if args.commit:
        session.commit()
    session.close()


if __name__ == "__main__":
    main()
