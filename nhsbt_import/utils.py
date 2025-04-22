"""
This module contains a number of utility functions that are used by the main
nhsbt_import.py script.

Functions:
    add_df_row(df, row): Adds a row to a dataframe
    args_parse(argv): Preforms some check on the inputs from the command line
    check_missing_patients(session, file_data): Checks for patients missing from the file
    check_missing_transplants(session, file_data): Checks for transplants missing from the file
    compare_patients(incoming_patient, existing_patient): Compares incoming and existing patient data
    compare_transplants(incoming_transplant, existing_transplant): Compares incoming and existing transplant data
    create_df(name, columns): Creates a dataframe
    create_incoming_patient(index, row): Creates an incoming patient object
    create_incoming_transplant(row, transplant_counter): Creates an incoming transplant object
    create_logs(directory): Creates a logger
    create_output_dfs(df_columns): Creates all the output dataframes
    create_session(): Creates a database session
    deleted_patient_check(session, file_patients): Checks patient identifiers against the deleted patient table
    format_bool(value): Converts a value to a bool
    format_date(str_date): Converts a string to a date. Returns None if the string is empty
    format_int(value): Converts a value to an int
    format_str(value): Converts a value to a string
    get_input_file_path(directory): Checks the supplied directory for the NHSBT
    make_deleted_patient_row(match_type, deleted_patient): Creates a row for the deleted patient sheet
    make_missing_patient_row(match_type, missing_patient): Creates a row for the missing patient sheet
    make_missing_transplant_match_row(missing_transplant): Creates a row for the missing transplant sheet
    make_patient_match_row(match_type, incoming_patient, existing_patient): Creates a row for the patient match sheet
    make_transplant_match_row(match_type, incoming_transplant, existing_transplant): Creates a row for the transplant match sheet
    update_nhsbt_patient(incoming_patient, existing_patient): Updates an existing patient
    update_nhsbt_transplant(incoming_transplant, existing_transplant): Updates an existing transplant
    nhsbt_clean(unclean_dataframe): Cleans up the dataframe
"""

import argparse
import datetime
import logging
import logging.config
import os
import sys
import re
import csv
from typing import Optional, Union, Any

import nhs_number  # type:ignore
from nhs_number import NhsNumber  # type:ignore
from tqdm import tqdm

import pandas as pd
from dateutil.parser import parse
from openpyxl import Workbook
from openpyxl.styles.fills import PatternFill
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from ukrr_models.nhsbt_models import UKTPatient, UKTTransplant  # type: ignore
from ukrr_models.rr_models import UKRR_Deleted_Patient  # type: ignore

log = logging.getLogger(__name__)


def add_df_row(df: pd.DataFrame, row: dict[str, str]) -> pd.DataFrame:
    """
    Adds a row to a dataframe

    Args:
        df (pd.DataFrame): dataframe to add row to
        row (dict[str, str]): row to add

    Returns:
        pd.DataFrame: dataframe with row added
    """
    row_df = pd.DataFrame(row, index=[0])
    return pd.concat([df, row_df], ignore_index=True)


def args_parse(argv=None) -> argparse.Namespace:
    """
    Preforms some check on the inputs. Firstly, if no input are provided the
    help text is displayed. Then a check is done to make sure that the
    provided input resolves to a valid path and that the path it resolves to
    is a directory not a file.

    The argv here are used for testing purposes

    Args:
        argv (list, optional): List of inputs. Defaults to None.

    Raises:
        NotADirectoryError: If path not found
        NotADirectoryError: If path does not resolve to a directory

    Returns:
        argparse.Namespace:
    """
    parser = argparse.ArgumentParser(description="nhsbt_import")
    parser.add_argument(
        "-d",
        "--directory",
        type=str,
        help="Specify the directory that holds the input file",
    )
    parser.add_argument(
        "--commit",
        action="store_true",
        help="Flag to turn on committing to the database",
    )

    args = parser.parse_args(argv)

    if len(sys.argv) <= 1:
        print(parser.format_help())
        sys.exit(1)

    if not os.path.exists(args.directory):
        raise NotADirectoryError(f"{args.directory} not found")

    if not os.path.isdir(args.directory):
        raise NotADirectoryError(f"Path is not a directory: {args.directory}")

    return args


def check_missing_patients(session: Session, file_data: list[int]) -> list[int]:
    """
    Checks for patients missing from the file

    Args:
        session (Session): a database session
        file_data (list[int]): a list of patient identifiers from the file

    Returns:
        list[int]: a list of patient identifiers missing from the file
    """
    results = session.query(UKTPatient.uktssa_no).all()
    db_data = [result[0] for result in results]

    return list(set(db_data) - set(file_data))


def check_missing_transplants(session: Session, file_data: list[str]) -> list[str]:
    """
    Checks for transplants missing from the file

    Args:
        session (Session): a database session
        file_data (list[str]): a list of transplant identifiers from the file

    Returns:
        list[str]: a list of transplant identifiers missing from the file
    """
    results = session.query(UKTTransplant.registration_id).all()
    db_data = [result[0] for result in results]

    return list(set(db_data) - set(file_data))


def clean_cell_value(cell_value):
    if isinstance(cell_value, str):
        return re.sub(r"[^\x00-\x7F]", "", cell_value.replace("\x00", ""))
    return cell_value


def clean_csv(input_filename):
    with open(input_filename, newline="", encoding="utf-8", errors="replace") as infile:
        reader = csv.reader(infile)
        rows = list(reader)

        cleaned_rows = []
        for row in tqdm(rows, desc="Cleaning null bytes and ASCII"):
            cleaned_row = [clean_cell_value(cell) for cell in row]
            cleaned_rows.append(cleaned_row)

    with open(input_filename, "w", newline="", encoding="utf-8") as outfile:
        writer = csv.writer(outfile)
        for cleaned_row in tqdm(cleaned_rows, desc="Writing rows"):
            writer.writerow(cleaned_row)


def compare_patients(
    incoming_patient: UKTPatient, existing_patient: UKTPatient
) -> bool:
    """
    Compares incoming and existing patient data. Ignore rr_no as it will never match
    because it is always None in the incoming data.

    Args:
        incoming_patient (UKTPatient): An incoming patient object
        existing_patient (UKTPatient): An existing patient object

    Returns:
        bool: True if the data matches, False otherwise
    """
    # Ignore rr_no as it will never match
    if incoming_patient.surname != existing_patient.surname:
        return False
    if incoming_patient.forename != existing_patient.forename:
        return False
    if incoming_patient.sex != existing_patient.sex:
        return False
    # if incoming_patient.post_code != existing_patient.post_code:
    #     return False
    if incoming_patient.new_nhs_no != existing_patient.new_nhs_no:
        return False
    if incoming_patient.chi_no != existing_patient.chi_no:
        return False
    if incoming_patient.hsc_no != existing_patient.hsc_no:
        return False
    if incoming_patient.ukt_date_death != existing_patient.ukt_date_death:
        return False
    return incoming_patient.ukt_date_birth == existing_patient.ukt_date_birth


def compare_transplants(
    incoming_transplant: UKTTransplant, existing_transplant: UKTTransplant
) -> bool:
    """
    Compares incoming and existing transplant data.

    Args:
        incoming_transplant (UKTPatient): An incoming transplant object
        existing_transplant (UKTPatient): An existing transplant object

    Returns:
        bool: True if the data matches, False otherwise
    """
    if incoming_transplant.transplant_id != existing_transplant.transplant_id:
        return False
    if incoming_transplant.uktssa_no != existing_transplant.uktssa_no:
        return False
    if incoming_transplant.transplant_date != existing_transplant.transplant_date:
        return False
    if incoming_transplant.transplant_type != existing_transplant.transplant_type:
        return False
    if incoming_transplant.transplant_organ != existing_transplant.transplant_organ:
        return False
    if incoming_transplant.transplant_unit != existing_transplant.transplant_unit:
        return False
    if incoming_transplant.ukt_fail_date != existing_transplant.ukt_fail_date:
        return False
    if incoming_transplant.registration_id != existing_transplant.registration_id:
        return False
    if incoming_transplant.registration_date != existing_transplant.registration_date:
        return False
    if (
        incoming_transplant.registration_date_type
        != existing_transplant.registration_date_type
    ):
        return False
    if (
        incoming_transplant.registration_end_date
        != existing_transplant.registration_end_date
    ):
        return False
    if (
        incoming_transplant.registration_end_status
        != existing_transplant.registration_end_status
    ):
        return False
    if (
        incoming_transplant.transplant_consideration
        != existing_transplant.transplant_consideration
    ):
        return False
    if (
        incoming_transplant.transplant_dialysis
        != existing_transplant.transplant_dialysis
    ):
        return False
    if (
        incoming_transplant.transplant_relationship
        != existing_transplant.transplant_relationship
    ):
        return False
    if incoming_transplant.transplant_sex != existing_transplant.transplant_sex:
        return False
    if incoming_transplant.cause_of_failure != existing_transplant.cause_of_failure:
        return False
    if (
        incoming_transplant.cause_of_failure_text
        != existing_transplant.cause_of_failure_text
    ):
        return False
    if incoming_transplant.cit_mins != existing_transplant.cit_mins:
        return False
    if incoming_transplant.hla_mismatch != existing_transplant.hla_mismatch:
        return False
    return incoming_transplant.ukt_suspension == existing_transplant.ukt_suspension


def colour_differences(wb: Workbook, sheet_name: str):
    """
    Highlights the difference between two cells that are next to each other
    in a sheet. Highlight is light blue.

    Args:
        wb (Workbook): A excel spreadsheet
        sheet_name (Worksheet): A sheet from the spreadsheet
    """
    sheet = wb[sheet_name]
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if differences := find_differences(row):
            for first_column, second_column in differences.items():
                fill = PatternFill(
                    start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
                )
                sheet.cell(row=row_number, column=first_column).fill = fill
                sheet.cell(row=row_number, column=second_column).fill = fill


def column_is_int(df: pd.DataFrame, column: str):
    """
    Check to see if everything in a dataframe column is an int

    Args:
        df (pd.DataFrame): A dataframe
        column (str):

    Raises:
        ValueError: Raised if something other than int detected
    """
    df[column] = pd.to_numeric(df[column], errors="coerce")
    if df[column].isna().any():
        log.error("UKTR ID column contains blanks or non-numbers")
        raise ValueError("UKTR ID column contains blanks or non-numbers")


def create_df(name: str, columns: dict[str, list[str]]) -> pd.DataFrame:
    """
    Creates a dataframe

    Args:
        name (str): Name of the dataframe
        columns (dict[str, list[str]]): Columns for the dataframe

    Returns:
        pd.DataFrame: A dataframe with the supplied columns and given name
    """
    return pd.DataFrame(columns=columns[name])


def validate_and_correct_nhs_numbers(row,row_index)-> pd.Series:
    """
   Validates and corrects the NHS numbers in the provided row of data.
    """
    invalids = validate_numbers(row)

    if invalids and any([i[1] for i in invalids]):
        swappable = {
            "UKTR_RNHS_NO": None,
            "UKTR_RCHI_NO_NI": None,
            "UKTR_RCHI_NO_SCOT": None,
        }
        for index, number, region, original_number in invalids:
            if number is None:
                continue
            match number.region:
                case nhs_number.REGION_ENGLAND:
                    old_index = "UKTR_RNHS_NO"
                case nhs_number.REGION_NORTHERN_IRELAND:
                    old_index = "UKTR_RCHI_NO_NI"
                case nhs_number.REGION_SCOTLAND:
                    old_index = "UKTR_RCHI_NO_SCOT"
                case _:
                    message = f"invalid number provided and can not be converted to region, check row {row_index}"
                    log.error(message)
                    raise ValueError(message)

            swappable[old_index] = original_number  # this avoids int conversions

        for index, value in swappable.items():
            if value:
                value = int(value)
            row[index] = value

        invalids = validate_numbers(row)
        if any([i[1] for i in invalids]):
            message = f"invalid number provided must be a valid number, check row {row_index}"
            log.error(message)
            raise ValueError(message)

    return row


def validate_numbers(row):
    nhs_no = NhsNumber(str(row["UKTR_RNHS_NO"])) if row["UKTR_RNHS_NO"] else None
    chi_no = (
        NhsNumber(str(row["UKTR_RCHI_NO_SCOT"])) if row["UKTR_RCHI_NO_SCOT"] else None
    )
    hsc_no = NhsNumber(str(row["UKTR_RCHI_NO_NI"])) if row["UKTR_RCHI_NO_NI"] else None
    validations = [
        ("UKTR_RNHS_NO", nhs_no, nhs_number.REGION_ENGLAND, row["UKTR_RNHS_NO"]),
        (
            "UKTR_RCHI_NO_NI",
            hsc_no,
            nhs_number.REGION_NORTHERN_IRELAND,
            row["UKTR_RCHI_NO_NI"],
        ),
        (
            "UKTR_RCHI_NO_SCOT",
            chi_no,
            nhs_number.REGION_SCOTLAND,
            row["UKTR_RCHI_NO_SCOT"],
        ),
    ]
    invalids = []
    for index, number, region, original_value in validations:
        if number is None:
            invalids.append((index, number, region, original_value))
        elif number.region != region:
            invalids.append((index, number, region, original_value))
        else:
            pass
    return invalids


def create_incoming_patient(index: int, row: pd.Series) -> UKTPatient:
    """
    Creates an incoming patient object

    Args:
        index (int): A row index
        row (pd.Series): The row data to create the patient from

    Raises:
        ValueError: Raised if UKTR_ID is not a valid number

    Returns:
        UKTPatient: A UKTPatient object containing the patient data
    """
    uktssa_no = format_int(row["UKTR_ID"])
    if uktssa_no == 0 or not isinstance(uktssa_no, int):
        message = f"UKTR_ID must be a valid number, check row {index + 1}"
        log.error(message)
        raise ValueError(message)

    row = validate_and_correct_nhs_numbers(row,index+1)
    if postcode := format_postcode(row["UKTR_RPOSTCODE"]):
        if len(postcode) < 2 or len(postcode) > 8:
            log.warning("Postcode length error on row %s: %s", index, postcode)

        if not postcode[:1].isalpha():
            log.warning("Incorrect postcode format on row %s: %s", index, postcode)

    return UKTPatient(
        uktssa_no=uktssa_no,
        surname=format_str(row["UKTR_RSURNAME"]),
        forename=format_str(row["UKTR_RFORENAME"]),
        sex=format_sex(row["UKTR_RSEX"], index),
        post_code=postcode,
        new_nhs_no=format_int(row["UKTR_RNHS_NO"]),
        chi_no=format_int(row["UKTR_RCHI_NO_SCOT"]),
        hsc_no=format_int(row["UKTR_RCHI_NO_NI"]),
        rr_no=None,
        ukt_date_death=format_date(row["UKTR_DDATE"]),
        ukt_date_birth=format_date(row["UKTR_RDOB"]),
    )


def create_incoming_transplant(
    index: int, row: pd.Series, transplant_counter: int
) -> UKTTransplant:
    """
    Creates an incoming transplant object. Transplant_counter is used to identify
    the transplant as there can be more than one transplant per patient.

    Args:
        row (pd.Series): The row data to create the transplant from
        transplant_counter (int): A counter to identify the transplant

    Returns:
        UKTTransplant: A UKTTransplant object containing the transplant data
    """
    tx_unit = format_str(row[f"uktr_tx_unit{transplant_counter}"])

    return UKTTransplant(
        transplant_id=format_int(row[f"uktr_tx_id{transplant_counter}"]),
        uktssa_no=format_int(row["UKTR_ID"]),
        transplant_date=format_date(row[f"uktr_txdate{transplant_counter}"]),
        transplant_type=format_str(row[f"uktr_dgrp{transplant_counter}"]),
        transplant_organ=format_str(row[f"uktr_tx_type{transplant_counter}"]),
        transplant_unit=None if tx_unit == "" else tx_unit,
        ukt_fail_date=format_date(row[f"uktr_faildate{transplant_counter}"]),
        rr_no=None,
        registration_id=f'{format_int(row["UKTR_ID"])}_{transplant_counter}',
        registration_date=format_date(row[f"uktr_date_on{transplant_counter}"]),
        registration_date_type=format_str(row[f"uktr_list_status{transplant_counter}"]),
        registration_end_date=format_date(
            row[f"uktr_removal_date{transplant_counter}"]
        ),
        registration_end_status=format_str(row[f"uktr_endstat{transplant_counter}"]),
        transplant_consideration=format_str(row[f"uktr_tx_list{transplant_counter}"]),
        transplant_dialysis=format_str(row[f"uktr_dial_at_tx{transplant_counter}"]),
        transplant_relationship=format_str(
            row[f"uktr_relationship{transplant_counter}"]
        ),
        transplant_sex=format_sex(row[f"uktr_dsex{transplant_counter}"], index),
        cause_of_failure=format_str(format_int(row[f"uktr_cof{transplant_counter}"])),
        cause_of_failure_text=format_str(
            row[f"uktr_other_cof_text{transplant_counter}"]
        ),
        cit_mins=format_str(row[f"uktr_cit_mins{transplant_counter}"]),
        hla_mismatch=format_str(row[f"uktr_hla_mm{transplant_counter}"]),
        ukt_suspension=format_bool(row[f"uktr_suspension_{transplant_counter}"]),
    )


def create_logs(directory: str) -> logging.Logger:
    """
    Uses the supplied directory for the NHSBT file to set up a errors file.
    Anything with a log level of warning or higher is written.

    Args:
        directory (str): The directory that holds the input file

    Returns:
        logging.Logger: A logger nhsbt_logger
    """
    errors_file_path = os.path.abspath(f"{directory}/errors.log").replace("\\", "/")

    logging.config.fileConfig(
        fname="logconf.conf",
        disable_existing_loggers=False,
        defaults={"log_file_name": errors_file_path},
    )
    return logging.getLogger("nhsbt_import")


def create_output_dfs(df_columns: dict[str, list[str]]) -> dict[str, pd.DataFrame]:
    """
    Creates all the output dataframes that are latter saved as an excel file. Also does
    some type conversions for bool columns to get round an issue with columns that
    have blank cells.

    Args:
        df_columns (dict[str, list[str]]): Includes all sheet names and columns

    Returns:
        dict[str, pd.DataFrame]: All the output dataframes
    """
    output_dfs = {df: create_df(df, df_columns) for df in df_columns}
    output_dfs["new_transplants"]["UKT Suspension - NHSBT"] = output_dfs[
        "new_transplants"
    ]["UKT Suspension - NHSBT"].astype(bool)

    output_dfs["updated_transplants"]["UKT Suspension - NHSBT"] = output_dfs[
        "updated_transplants"
    ]["UKT Suspension - NHSBT"].astype(bool)

    output_dfs["updated_transplants"]["UKT Suspension - RR"] = output_dfs[
        "updated_transplants"
    ]["UKT Suspension - RR"].astype(bool)

    return output_dfs


def create_session() -> Session:
    """
    Creates a database session

    Returns:
        Session: A database session
    """
    driver = "SQL+Server+Native+Client+11.0"
    engine = create_engine(f"mssql+pyodbc://rr-sql-live/renalreg?driver={driver}")
    # engine = create_engine("postgresql://postgres:password@localhost:5432/radar")

    return Session(engine, future=True)


def deleted_patient_check(session: Session, file_patients: list[str]) -> list[str]:
    """
    Checks patient identifiers against the deleted patient table

    Args:
        session (Session): a database session
        file_patients (list[str]): a list of patient identifiers from the file

    Returns:
        list[str]: a list of patient identifiers that have been deleted
    """
    results = session.query(UKRR_Deleted_Patient.uktssa_no).all()
    db_patients = {result[0] for result in results}

    return list(db_patients.intersection(set(file_patients)))


def find_differences(row: tuple):
    """
    Compares two cells to see if they are different

    Args:
        row (tuple): A row in a spreadsheet

    Returns:
        Dict: The differing values
    """

    sliced_row = row[3:]
    last_index = len(sliced_row) - 1
    differences = {}
    while last_index > 0:
        first_index, second_index = last_index, last_index - 1
        if str(sliced_row[first_index]) != str(sliced_row[second_index]):
            differences[first_index + 4] = second_index + 4
        last_index -= 2

    return differences


def format_bool(value: Any) -> Optional[bool]:
    """
    Converts a value to a bool

    Args:
        value (Any): A value to convert

    Returns:
        Optional[bool]: A bool or None
    """
    if value in ("0", "0.0", 0, 0.0, "False", "false", False):
        return False
    return True if value in ("1", "1.0", 1, 1.0, "True", "true", True) else None


def format_date(
    str_date: Any, strip_time=False
) -> Optional[Union[datetime.datetime, datetime.date]]:
    """
    Converts a string to a datetime. Returns None if the string is empty

    Args:
        str_date (Optional[str]): A string to convert

    Returns:
        Optional[date]: A date or None
    """
    if not str_date or pd.isna(str_date):
        return None

    if isinstance(str_date, datetime.datetime):
        return str_date.date() if strip_time else str_date

    if isinstance(str_date, datetime.date):
        return (
            str_date
            if strip_time
            else datetime.datetime.combine(str_date, datetime.time.min)
        )

    if str_date[:4].isdigit():
        try:
            parsed_date = parse(str_date, yearfirst=True)
        except (ValueError, TypeError):
            log.warning("%s is not a valid date", str_date)
            return None
    else:
        try:
            parsed_date = parse(str_date, dayfirst=True)
        except (ValueError, TypeError):
            log.warning("%s is not a valid date", str_date)
            return None

    return parsed_date.date() if strip_time else parsed_date


def format_int(value: Any) -> Optional[int]:
    """
    Converts a value to an int. Deals with NaNs

    Args:
        value (Any): A value to convert

    Returns:
        Optional[int]: An int or None
    """
    try:
        return None if pd.isna(value) else int(value)
    except (ValueError, TypeError):
        return None


def format_sex(value: Any, index: int) -> Optional[str]:
    """
    Attempts to convert a value to a recognised NHS gender code

    Args:
        value (Any): value to be formatted
        index (int): row number in case reference is need for logs

    Returns:
        Optional[int]: Returns None if value can't be converted
    """
    message = f"Unrecognised sex at row {index + 1}"

    if not (formatted_value := format_str(value)):
        log.warning(message)
        return None
    if formatted_value in ("0", "1", "2", "9"):
        return formatted_value
    if formatted_value.lower() in ("not known", "not_known", "nk"):
        return "0"
    if formatted_value.lower() in ("male", "m", "1.0"):
        return "1"
    if formatted_value.lower() in ("female", "f", "2.0"):
        return "2"
    if formatted_value.lower() in ("not specified", "not_specified", "ns", "9.0"):
        return "9"

    log.warning(message)
    return None


def format_str(value: Any) -> Optional[str]:
    """
    Converts a value to a string. Deals with NaNs

    Args:
        value (Any): A value to convert

    Returns:
        str: A string or None
    """
    try:
        return None if pd.isna(value) else str(value)
    except (ValueError, TypeError):
        return None


def format_postcode(postcode: Optional[str]) -> Optional[str]:
    """
    Ensure that postcode is made up of two parts, second part is made
    up of exactly three characters.

    Args:
        postcode (Any): a string representing post code

    Returns:
        Optional[str]: formatted postcode
    """

    if (postcode := format_str(postcode)) is None:
        return postcode

    postcode = postcode.upper().strip()
    postcode = " ".join(postcode.split())

    if len(postcode.split()) > 2:
        postcode = "".join(postcode.split())

    if len(postcode.split()) == 1 and len(postcode) > 4 and len(postcode) < 8:
        postcode = f"{postcode[:-3]} {postcode[-3:]}"

    return postcode


def get_input_file_path(directory: str) -> str:
    """
    Checks the supplied directory for the NHSBT. Looks for CSV files
    in the directory and checks that there is only one. More than one
    raises an error

    Args:
        directory (str): Supplied input directory path

    Raises:
        ValueError: Raised if there is more than one CSV found

    Returns:
        str: File path
    """
    csv_list = [file for file in os.listdir(directory) if file.endswith(".csv")]

    if len(csv_list) > 1:
        log.warning(
            "Expected to find one import CSV file in %s, but found %s files.",
            directory,
            len(csv_list),
        )
        raise ValueError(
            f"Expected to find one import CSV file in {directory}, but found {len(csv_list)} files."
        )

    return os.path.join(directory, csv_list[0])


def make_deleted_patient_row(
    match_type: str, deleted_patient: UKRR_Deleted_Patient
) -> dict[str, str]:
    """
    Creates a row for the deleted patient sheet

    Args:
        match_type (str): The type of match
        deleted_patient (UKRR_Deleted_Patient): A deleted patient object

    Returns:
        dict[str, str]: A row for the deleted patient sheet
    """
    # TODO: [NHSBT-8] Add the other columns CHI etc
    return {
        "Match Type": match_type,
        "UKTSSA_No": deleted_patient.uktssa_no,
        "RR_No": deleted_patient.rr_no,
        "Surname - RR": deleted_patient.surname,
        "Forename - RR": deleted_patient.forename,
        "Sex - RR": deleted_patient.sex,
        "Date Birth - RR": str(
            format_date(deleted_patient.date_birth, strip_time=True)
        ),
        "NHS Number - RR": deleted_patient.nhs_no,
        "CHI Number - NHSBT": deleted_patient.chi_no,
        "HSC Number - NHSBT": deleted_patient.hsc_no,
    }


def make_missing_patient_row(
    match_type: str, missing_patient: UKTPatient
) -> dict[str, str]:
    """
    Creates a row for the missing patient sheet

    Args:
        match_type (str): The type of match
        missing_patient (UKTPatient): A missing patient object

    Returns:
        dict[str, str]: A row for the missing patient sheet
    """
    # TODO: [NHSBT-8] Add the other columns CHI etc
    return {
        "Match Type": match_type,
        "UKTSSA_No": missing_patient.uktssa_no,
        "Surname - RR": missing_patient.surname,
        "Forename - RR": missing_patient.forename,
        "Sex - RR": missing_patient.sex,
        "Date Birth - RR": str(
            format_date(missing_patient.ukt_date_birth, strip_time=True)
        ),
        "NHS Number - RR": missing_patient.new_nhs_no,
        "CHI Number - NHSBT": missing_patient.chi_no,
        "HSC Number - NHSBT": missing_patient.hsc_no,
    }


def make_missing_transplant_match_row(
    missing_transplant: UKTTransplant,
) -> dict[str, str | int | bool | None]:
    """
    Creates a row for the missing transplant sheet

    Args:
        missing_transplant (UKTTransplant): A missing transplant object

    Returns:
        dict[str, str | int | bool | None]: A row for the missing transplant sheet
    """
    return {
        "Match Type": "Missing",
        "UKTSSA_No": missing_transplant.uktssa_no,
        "Transplant ID - RR": missing_transplant.transplant_id,
        "Registration ID - RR": missing_transplant.registration_id,
        "Transplant Date - RR": str(
            format_date(missing_transplant.transplant_date, strip_time=True)
        ),
        "Transplant Type - RR": missing_transplant.transplant_type,
        "Transplant Organ - RR": missing_transplant.transplant_organ,
        "Transplant Unit - RR": missing_transplant.transplant_unit,
        "Registration Date - RR": str(
            format_date(missing_transplant.registration_date, strip_time=True)
        ),
        "Registration Date Type - RR": missing_transplant.registration_date_type,
        "Registration End Date - RR": str(
            format_date(missing_transplant.registration_end_date, strip_time=True)
        ),
        "Registration End Status - RR": missing_transplant.registration_end_status,
        "Transplant Consideration - RR": missing_transplant.transplant_consideration,
        "Transplant Dialysis - RR": missing_transplant.transplant_dialysis,
        "Transplant Relationship - RR": missing_transplant.transplant_relationship,
        "Transplant Sex - RR": missing_transplant.transplant_sex,
        "Cause of Failure - RR": missing_transplant.cause_of_failure,
        "Cause of Failure Text - RR": missing_transplant.cause_of_failure_text,
        "CIT Mins - RR": missing_transplant.cit_mins,
        "HLA Mismatch - RR": missing_transplant.hla_mismatch,
        "UKT Suspension - RR": format_bool(missing_transplant.ukt_suspension),
    }


def make_patient_match_row(
    match_type: str,
    incoming_patient: UKTPatient,
    existing_patient: Optional[UKTPatient],
) -> dict[str, str]:
    """
    Creates a row for the patient match sheet

    Args:
        match_type (str): The type of match
        incoming_patient (UKTPatient): An incoming patient object
        existing_patient (Optional[UKTPatient]): An existing patient object

    Returns:
        dict[str, str]: A row for the patient match sheet
    """
    # TODO: [NHSBT-8] Add the other columns CHI etc
    patient_row = {
        "Match Type": match_type,
        "UKTSSA_No": incoming_patient.uktssa_no,
        "Surname - NHSBT": incoming_patient.surname,
        "Forename - NHSBT": incoming_patient.forename,
        "Sex - NHSBT": incoming_patient.sex,
        "Date Birth - NHSBT": format_date(
            incoming_patient.ukt_date_birth, strip_time=True
        ),
        "Date Death - NHSBT": format_date(
            incoming_patient.ukt_date_death, strip_time=True
        ),
        "NHS Number - NHSBT": incoming_patient.new_nhs_no,
        "CHI Number - NHSBT": incoming_patient.chi_no,
        "HSC Number - NHSBT": incoming_patient.hsc_no,
        "Postcode - NHSBT": incoming_patient.post_code,
    }

    if existing_patient:
        patient_row["RR_No"] = existing_patient.rr_no
        patient_row["Surname - RR"] = existing_patient.surname
        patient_row["Forename - RR"] = existing_patient.forename
        patient_row["Sex - RR"] = existing_patient.sex
        patient_row["Date Birth - RR"] = format_date(
            existing_patient.ukt_date_birth, strip_time=True
        )
        patient_row["Date Death - RR"] = format_date(
            existing_patient.ukt_date_death, strip_time=True
        )
        patient_row["NHS Number - RR"] = existing_patient.new_nhs_no
        patient_row["CHI Number - RR"] = existing_patient.chi_no
        patient_row["HSC Number - RR"] = existing_patient.hsc_no
        patient_row["Postcode - RR"] = existing_patient.post_code

    return patient_row


def make_transplant_match_row(
    match_type: str,
    incoming_transplant: UKTTransplant,
    existing_transplant: Optional[UKTTransplant],
) -> dict[str, str]:
    """
    Creates a row for the transplant match sheet

    Args:
        match_type (str): The type of match
        incoming_transplant (UKTTransplant): An incoming transplant object
        existing_transplant (Optional[UKTTransplant]): An existing transplant object

    Returns:
        dict[str, str]: A row for the transplant match sheet
    """
    transplant_row = {
        "Match Type": match_type,
        "UKTSSA_No": incoming_transplant.uktssa_no,
        "Transplant ID - NHSBT": incoming_transplant.transplant_id,
        "Transplant Date - NHSBT": format_date(
            incoming_transplant.transplant_date, strip_time=True
        ),
        "Transplant Type - NHSBT": incoming_transplant.transplant_type,
        "Transplant Organ - NHSBT": incoming_transplant.transplant_organ,
        "Transplant Unit - NHSBT": incoming_transplant.transplant_unit,
        "Registration Date - NHSBT": format_date(
            incoming_transplant.registration_date, strip_time=True
        ),
        "Registration Date Type - NHSBT": incoming_transplant.registration_date_type,
        "Registration End Date - NHSBT": format_date(
            incoming_transplant.registration_end_date, strip_time=True
        ),
        "Registration End Status - NHSBT": incoming_transplant.registration_end_status,
        "Transplant Consideration - NHSBT": incoming_transplant.transplant_consideration,
        "Transplant Dialysis - NHSBT": incoming_transplant.transplant_dialysis,
        "Transplant Relationship - NHSBT": incoming_transplant.transplant_relationship,
        "Transplant Sex - NHSBT": incoming_transplant.transplant_sex,
        "Cause of Failure - NHSBT": incoming_transplant.cause_of_failure,
        "Cause of Failure Text - NHSBT": incoming_transplant.cause_of_failure_text,
        "CIT Mins - NHSBT": incoming_transplant.cit_mins,
        "HLA Mismatch - NHSBT": incoming_transplant.hla_mismatch,
        "UKT Suspension - NHSBT": incoming_transplant.ukt_suspension,
    }

    if existing_transplant:
        transplant_row["Transplant ID - RR"] = existing_transplant.transplant_id
        transplant_row["Registration ID - RR"] = existing_transplant.registration_id
        transplant_row["Transplant Date - RR"] = (
            format_date(existing_transplant.transplant_date, strip_time=True),
        )
        transplant_row["Transplant Type - RR"] = existing_transplant.transplant_type
        transplant_row["Transplant Organ - RR"] = existing_transplant.transplant_organ
        transplant_row["Transplant Unit - RR"] = existing_transplant.transplant_unit
        transplant_row["Registration Date - RR"] = (
            format_date(existing_transplant.registration_date, strip_time=True),
        )
        transplant_row["Registration Date Type - RR"] = (
            existing_transplant.registration_date_type
        )
        transplant_row["Registration End Date - RR"] = (
            format_date(existing_transplant.registration_end_date, strip_time=True),
        )
        transplant_row["Registration End Status - RR"] = (
            existing_transplant.registration_end_status
        )
        transplant_row["Transplant Consideration - RR"] = (
            existing_transplant.transplant_consideration
        )
        transplant_row["Transplant Dialysis - RR"] = (
            existing_transplant.transplant_dialysis
        )
        transplant_row["Transplant Relationship - RR"] = (
            existing_transplant.transplant_relationship
        )
        transplant_row["Transplant Sex - RR"] = existing_transplant.transplant_sex
        transplant_row["Cause of Failure - RR"] = existing_transplant.cause_of_failure
        transplant_row["Cause of Failure Text - RR"] = (
            existing_transplant.cause_of_failure_text
        )
        transplant_row["CIT Mins - RR"] = existing_transplant.cit_mins
        transplant_row["HLA Mismatch - RR"] = existing_transplant.hla_mismatch
        transplant_row["UKT Suspension - RR"] = existing_transplant.ukt_suspension

    return transplant_row


def update_nhsbt_patient(
    incoming_patient: UKTPatient, existing_patient: UKTPatient
) -> UKTPatient:
    """
    Updates an existing patient with incoming patient data. Incoming RR will always be
    None so preserve existing

    Args:
        incoming_patient (UKTPatient): An incoming patient object
        existing_patient (UKTPatient): An existing patient object

    Returns:
        UKTPatient: An updated patient object
    """
    # Incoming RR will always be None so preserve existing
    existing_patient.uktssa_no = incoming_patient.uktssa_no
    existing_patient.surname = incoming_patient.surname
    existing_patient.forename = incoming_patient.forename
    existing_patient.sex = incoming_patient.sex
    existing_patient.post_code = incoming_patient.post_code
    existing_patient.new_nhs_no = incoming_patient.new_nhs_no
    existing_patient.chi_no = incoming_patient.chi_no
    existing_patient.hsc_no = incoming_patient.hsc_no
    existing_patient.ukt_date_death = incoming_patient.ukt_date_death
    existing_patient.ukt_date_birth = incoming_patient.ukt_date_birth


def update_nhsbt_transplant(
    incoming_transplant: UKTTransplant, existing_transplant: UKTTransplant
) -> UKTTransplant:
    """
    Updates an existing transplant with incoming transplant data. Incoming RR will
    always be None so preserve existing

    Args:
        incoming_transplant (UKTTransplant): An incoming transplant object
        existing_transplant (UKTTransplant): An existing transplant object

    Returns:
        UKTTransplant: An updated transplant object
    """
    # Incoming RR will always be None so preserve existing
    existing_transplant.transplant_id = incoming_transplant.transplant_id
    existing_transplant.uktssa_no = incoming_transplant.uktssa_no
    existing_transplant.transplant_date = incoming_transplant.transplant_date
    existing_transplant.transplant_type = incoming_transplant.transplant_type
    existing_transplant.transplant_organ = incoming_transplant.transplant_organ
    existing_transplant.transplant_unit = incoming_transplant.transplant_unit
    existing_transplant.ukt_fail_date = incoming_transplant.ukt_fail_date
    existing_transplant.registration_id = incoming_transplant.registration_id
    existing_transplant.registration_date = incoming_transplant.registration_date
    existing_transplant.registration_date_type = (
        incoming_transplant.registration_date_type
    )
    existing_transplant.registration_end_date = (
        incoming_transplant.registration_end_date
    )
    existing_transplant.registration_end_status = (
        incoming_transplant.registration_end_status
    )
    existing_transplant.transplant_consideration = (
        incoming_transplant.transplant_consideration
    )
    existing_transplant.transplant_dialysis = incoming_transplant.transplant_dialysis
    existing_transplant.transplant_relationship = (
        incoming_transplant.transplant_relationship
    )
    existing_transplant.transplant_sex = incoming_transplant.transplant_sex
    existing_transplant.cause_of_failure = incoming_transplant.cause_of_failure
    existing_transplant.cause_of_failure_text = (
        incoming_transplant.cause_of_failure_text
    )
    existing_transplant.cit_mins = incoming_transplant.cit_mins
    existing_transplant.hla_mismatch = incoming_transplant.hla_mismatch
    existing_transplant.ukt_suspension = incoming_transplant.ukt_suspension
